import openpyxl


def format_data(file_path, sheet_name, column_index):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
       # Iterate through the specified column and add quotes around each entry
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, values_only=True), start=2):
        for cell in row:
            if cell is not None:
                cell_value = str(cell)
                # this next line formats the cell to match the string within the f-string, in this case it adds quotes and a comma
                sheet.cell(row=row_num, column=column_index).value = f'"{cell_value}",'   

    wb.save(file_path)

file_path = "./data.xlsx"
sheet_name = "Contribution Totals"
column_index = 3  # Assuming you want to format the third column (C)

format_data(file_path, sheet_name, column_index)
    
